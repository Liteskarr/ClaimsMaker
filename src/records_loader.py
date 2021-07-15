from datetime import datetime
from decimal import Decimal
from functools import reduce
from itertools import count
from operator import or_
from typing import Callable

import openpyxl
from PyQt5.QtCore import (QObject,
                          pyqtSignal)
from openpyxl import load_workbook

from src.config import MainConfig
from src.record import Record

KEYWORDS = {
    'ДАТА': lambda r: {'Date': r},
    'НОМЕР': lambda r: {'Number': r},
    'СЧЕТ': lambda r: {'Account': r},
    'ИНН': lambda r: {'INN': r},
    'ИМЯ': lambda r: {'Name': r},
    'ДЕБЕТ': lambda r: {'Debit': Decimal(r).quantize(Decimal('1.00')) if r else Decimal(0)},
    'КРЕДИТ': lambda r: {'Credit': Decimal(r).quantize(Decimal('1.00')) if r else Decimal(0)},
    'ОСНОВАНИЕ': lambda r: {'Reason': r},
}

NoneRecord = Record(**reduce(or_, (func(None) for func in KEYWORDS.values()), {}))


def _find_labels(book: openpyxl.Workbook, sheet_name: str) -> dict[int, str]:
    result = {}
    sheet = book[sheet_name]
    for column in count(1):
        label = sheet.cell(row=MainConfig.ROW_LABEL, column=column).value
        if label in KEYWORDS:
            result[column] = label
        elif not bool(label):
            break
    return result


class RecordsLoader(QObject):
    record_read = pyqtSignal(str, int, int)  # filepath, number, count

    def __init__(self):
        super().__init__()
        self._records = []
        self._errors = []

    def apply_filter(self, func: Callable[[Record], bool]):
        self._records = list(filter(func, self._records))

    def get_records(self) -> list[Record]:
        return self._records

    def get_inns(self) -> list[str]:
        return list(set(map(lambda r: r.INN, self.get_records())))

    def get_errors(self) -> list[Record]:
        return self._errors

    def load_from_xlsx(self, filepath: str):
        book = load_workbook(filepath)
        sheet = book[MainConfig.SHEET_NAME]
        labels = _find_labels(book, MainConfig.SHEET_NAME)
        for r in range(MainConfig.ROW_DATA, sheet.max_row + 1):
            kwargs = {}
            try:
                for c in range(MainConfig.COLUMN_START, sheet.max_column + 1):
                    if c in labels.keys():
                        value = sheet.cell(row=r, column=c).value
                        if isinstance(value, datetime):
                            kwargs.update(KEYWORDS[labels[c]](value.strftime('%d.%m.%Y')))
                        else:
                            kwargs.update(KEYWORDS[labels[c]](value))
                if None not in set(kwargs.values()):
                    self._records.append(Record(**kwargs))
                    self.record_read.emit(filepath, r - 1, sheet.max_row)
            except:
                self._errors.append(kwargs)
